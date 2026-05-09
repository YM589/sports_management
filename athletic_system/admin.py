from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from .models import User, Event, Result, Team, Judge, Athlete, Announcement
from django import forms
from django.utils.html import format_html
from django.urls import path
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from django.http import HttpResponse, HttpResponseRedirect
# from reportlab.pdfgen import canvas
from io import BytesIO
from django.db.models import Count, Avg
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import io
import base64
from django.contrib.admin.sites import AdminSite
from django.core.exceptions import PermissionDenied
# from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
# from reportlab.lib.styles import getSampleStyleSheet
from django.db.models import Q
# from reportlab.lib import colors
# from reportlab.lib.pagesizes import letter
from django.utils import timezone
from django.contrib import messages
import matplotlib
matplotlib.use('Agg')
import openpyxl
import pandas as pd
import csv

class CustomAdminSite(AdminSite):
    def has_permission(self, request):
        """
        只允许管理员角色的用户访问后台
        """
        if not super().has_permission(request):
            return False
        return request.user.is_authenticated and request.user.role == 'admin'

    def logout(self, request, *args, **kwargs):
        # 调用默认的注销方法
        response = super().logout(request, *args, **kwargs)
        # 重定向到自定义的登录页面
        return redirect('logout')

# 创建自定义admin站点实例
admin_site = CustomAdminSite(name='custom_admin')

class AnnouncementForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 手动设置 priority 字段的初始值为 1
        self.fields['priority'].initial = 1

    class Meta:
        model = Announcement
        fields = '__all__'
        widgets = {
            'priority': forms.NumberInput(attrs={'min': 1,'max': 3, 'value': 1})
        }

    def clean_priority(self):
        priority = self.cleaned_data.get('priority')
        if priority is None:
            priority = 1  # 若未输入，设为默认值1
        if priority < 1 or priority > 3:
            raise forms.ValidationError('优先级必须在 1 到 3 之间')
        return priority

class AnnouncementAdmin(admin.ModelAdmin):
    form = AnnouncementForm
    list_display = ('title', 'content_summary', 'publish_time', 'priority', 'is_active')
    list_filter = ('priority', 'is_active', 'publish_time')
    search_fields = ('title', 'content')
    ordering = ('-priority', '-publish_time')

    def content_summary(self, obj):
        if len(obj.content) > 50:
            return obj.content[:50] + '...'
        return obj.content
    content_summary.short_description = '内容摘要'

admin_site.register(Announcement, AnnouncementAdmin)

class CustomUserAdmin(UserAdmin):
    list_display = ('username', 'email', 'role', 'phone', 'is_active')
    list_filter = ('role', 'is_active')
    search_fields = ('username', 'email', 'phone')
    
    # def has_delete_permission(self, request, obj=None):
    #     # 如果是单个对象的删除请求，返回False
    #     if obj is not None:
    #         return False
    #     # 对于批量删除操作（obj为None），返回True
    #     return True
    
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('个人信息', {'fields': ('first_name', 'last_name', 'email', 'phone')}),
        ('权限', {'fields': ('role', 'is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
    )
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2', 'role', 'email', 'phone'),
        }),
    )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if not request.user.is_superuser:
            qs = qs.filter(id=request.user.id)
        return qs

class JudgeAdmin(admin.ModelAdmin):
    list_display = ('user', 'title', 'specialty', 'get_events')
    search_fields = ('user__username', 'title', 'specialty')
    change_list_template = 'admin/judge_changelist.html'  # 添加自定义模板
    
    # def has_delete_permission(self, request, obj=None):
    #     # 如果是单个对象的删除请求，返回False
    #     if obj is not None:
    #         return False
    #     # 对于批量删除操作（obj为None），返回True
    #     return True
    
    def get_events(self, obj):
        events = Event.objects.filter(judges=obj.user)
        return ", ".join([event.name for event in events])
    get_events.short_description = '负责的比赛项目'
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if not request.user.is_superuser:
            qs = qs.filter(user=request.user)
        return qs

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-judges/', self.import_judges, name='import-judges'),
        ]
        return custom_urls + urls
    
    def import_judges(self, request):
        if request.method == 'POST':
            try:
                excel_file = request.FILES['excel_file']
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # 获取表头
                headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
                required_fields = ['username', 'password', 'email', 'title', 'specialty']
                
                # 验证必填字段
                missing_fields = [field for field in required_fields if field not in headers]
                if missing_fields:
                    self.message_user(
                        request, 
                        f'Excel文件缺少必填字段: {", ".join(missing_fields)}',
                        level=messages.ERROR
                    )
                    return redirect('..')
                
                # 获取字段索引
                field_indices = {field: headers.index(field) for field in required_fields}
                phone_index = headers.index('phone') if 'phone' in headers else None
                
                success_count = 0
                error_count = 0
                
                # 从第二行开始读取数据
                for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                    try:
                        # 创建用户
                        user = User.objects.create_user(
                            username=str(row[field_indices['username']].value),
                            password=str(row[field_indices['password']].value),
                            email=str(row[field_indices['email']].value),
                            role='judge',
                            phone=str(row[phone_index].value) if phone_index is not None and row[phone_index].value else ''
                        )
                        
                        # 创建裁判
                        Judge.objects.create(
                            user=user,
                            title=str(row[field_indices['title']].value),
                            specialty=str(row[field_indices['specialty']].value)
                        )
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        self.message_user(
                            request,
                            f'第{row_num}行导入失败: {str(e)}',
                            level=messages.ERROR
                        )
                
                self.message_user(
                    request,
                    f'导入完成。成功: {success_count}条, 失败: {error_count}条。',
                    level=messages.SUCCESS if error_count == 0 else messages.WARNING
                )
                
            except Exception as e:
                self.message_user(request, f'导入失败: {str(e)}', level=messages.ERROR)
            
            return redirect('..')
            
        return render(request, 'admin/import_judges.html', {
            'title': '导入裁判员',
            'opts': self.model._meta,
            'sample_format': '''
Excel文件格式说明：
第一行必须包含以下字段：
- username（用户名）*
- password（密码）*
- email（邮箱）*
- title（职称）*
- specialty（专长项目）*
- phone（电话，可选）

注：带*的为必填字段
            ''',
        })

class AthleteAdmin(admin.ModelAdmin):
    list_display = ('user', 'student_id', 'team', 'age', 'gender', 'get_events')
    list_filter = ('team', 'gender')
    search_fields = ('user__username', 'student_id')
    change_list_template = 'admin/athlete_changelist.html'  # 添加自定义模板
    
    # def has_delete_permission(self, request, obj=None):
    #     # 如果是单个对象的删除请求，返回False
    #     if obj is not None:
    #         return False
    #     # 对于批量删除操作（obj为None），返回True
    #     return True
    
    def get_events(self, obj):
        results = Result.objects.filter(athlete=obj.user)
        return ", ".join([f"{result.event.name}: {result.score or '未记录'}" for result in results])
    get_events.short_description = '参加的比赛及成绩'
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if not request.user.is_superuser:
            qs = qs.filter(user=request.user)
        return qs

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-athletes/', self.import_athletes, name='import-athletes'),
        ]
        return custom_urls + urls

    def import_athletes(self, request):
        if request.method == 'POST':
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            
            # 从第二行开始读取数据
            for row in ws.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                
                # 创建用户
                user = User.objects.create_user(
                    username=row_data['username'],
                    password=row_data['password'],
                    email=row_data['email'],
                    role='athlete',
                    phone=str(row_data.get('phone', ''))
                )
                
                # 获取或创建团队
                team, _ = Team.objects.get_or_create(name=row_data['team'])
                
                # 创建运动员
                Athlete.objects.create(
                    user=user,
                    student_id=str(row_data['student_id']),
                    team=team,
                    age=int(row_data['age']),
                    gender=row_data['gender']
                )
            
            self.message_user(request, '运动员导入成功')
            return redirect('..')
        return render(request, 'admin/import_athletes.html', {
            'title': '导入运动员',
            'opts': self.model._meta,
        })

class EventAdmin(admin.ModelAdmin):
    list_display = ('name', 'date', 'location', 'status', 'get_judges', 'get_participant_count', 'get_quit_count', 'registration_deadline')
    list_filter = ('status', 'date', 'location')
    search_fields = ('name', 'location', 'description')
    filter_horizontal = ('judges',)
    readonly_fields = ('get_participant_count', 'get_quit_count', 'get_registration_status')
    list_per_page = 20
    date_hierarchy = 'date'
    
    # def has_delete_permission(self, request, obj=None):
    #     # 如果是单个对象的删除请求，返回False
    #     if obj is not None:
    #         return False
    #     # 对于批量删除操作（obj为None），返回True
    #     return True

    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'description', 'max_participants', 'get_registration_status')
        }),
        ('时间地点', {
            'fields': ('date', 'location', 'registration_deadline')
        }),
        ('状态管理', {
            'fields': ('status',),
            'classes': ('wide',)
        }),
        ('裁判分配', {
            'fields': ('judges',),
            'classes': ('collapse',)
        }),
        ('统计信息', {
            'fields': ('get_participant_count', 'get_quit_count'),
            'classes': ('collapse',)
        })
    )
    
    def get_judges(self, obj):
        return format_html(
            '<br>'.join([f'{judge.username} ({judge.judge_profile.title})' for judge in obj.judges.all()])
        )
    get_judges.short_description = '负责裁判'
    
    def get_participant_count(self, obj):
        count = Result.objects.filter(event=obj).count()
        return format_html(
            '<span style="color: {}">{}/{}</span>',
            'red' if count >= obj.max_participants else 'green',
            count,
            obj.max_participants
        )
    get_participant_count.short_description = '参赛人数'
    
    def get_quit_count(self, obj):
        return Result.objects.filter(event=obj, is_quit=True).count()
    get_quit_count.short_description = '弃权人数'
    
    def get_registration_status(self, obj):
        if obj.registration_deadline:
            if obj.registration_deadline < timezone.now():
                return format_html('<span style="color: red">报名已截止</span>')
            else:
                return format_html('<span style="color: green">报名进行中</span>')
        return '未设置报名截止时间'
    get_registration_status.short_description = '报名状态'
    
    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('judges', 'judges__judge_profile')
    
    def save_model(self, request, obj, form, change):
        if not change:  # 新建项目时
            if obj.registration_deadline and obj.registration_deadline > obj.date:
                messages.error(request, '报名截止时间不能晚于比赛时间')
                return
            super().save_model(request, obj, form, change)
        else:  # 修改项目时
            old_obj = self.model.objects.get(pk=obj.pk)
            if old_obj.status != obj.status:
                if obj.status == 'ongoing':
                    if timezone.now() < obj.date:
                        messages.error(request, '比赛还未到开始时间，不能更改状态为进行中')
                        return
                elif obj.status == 'finished':
                    unconfirmed = Result.objects.filter(
                        event=obj,
                        confirmed=False,
                        is_quit=False
                    ).exists()
                    if unconfirmed:
                        messages.error(request, '还有未确认的成绩，不能结束比赛')
                        return
            super().save_model(request, obj, form, change)
    
    def response_change(self, request, obj):
        if "_manage-quits" in request.POST:
            return redirect('admin:athletic_system_result_changelist')
        return super().response_change(request, obj)
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_quit_button'] = True
        return super().change_view(
            request, object_id, form_url, extra_context=extra_context
        )

class ResultAdmin(admin.ModelAdmin):
    list_display = ('athlete', 'event', 'get_score_display', 'rank', 'get_status', 'created_at')
    list_filter = ('is_quit', 'confirmed', 'event__status', 'event__date')
    search_fields = ('athlete__username', 'event__name', 'quit_reason')
    ordering = ('event__date', 'rank')
    date_hierarchy = 'created_at'
    list_per_page = 20
    actions = ['export_as_csv', 'generate_statistics', 'mark_as_quit', 'remove_quit', 'confirm_selected']
    
    # def has_delete_permission(self, request, obj=None):
    #     # 如果是单个对象的删除请求，返回False
    #     if obj is not None:
    #         return False
    #     # 对于批量删除操作（obj为None），返回True
    #     return True
    
    fieldsets = (
        ('基本信息', {
            'fields': ('athlete', 'event')
        }),
        ('成绩信息', {
            'fields': ('score', 'rank', 'confirmed')
        }),
        ('弃权信息', {
            'fields': ('is_quit', 'quit_reason'),
            'classes': ('collapse',)
        })
    )
    
    def get_score_display(self, obj):
        if obj.is_quit:
            return format_html('<span style="color: red">已弃权</span>')
        elif obj.score is None:
            return format_html('<span style="color: gray">未记录</span>')
        else:
            return format_html('<b>{}</b>', obj.score)
    get_score_display.short_description = '成绩'
    
    def get_status(self, obj):
        if obj.is_quit:
            return format_html(
                '<span style="color: red" title="{}">已弃权</span>',
                obj.quit_reason or '无原因'
            )
        elif obj.confirmed:
            return format_html('<span style="color: green">已确认</span>')
        else:
            return format_html('<span style="color: orange">未确认</span>')
    get_status.short_description = '状态'
    
    def confirm_selected(self, request, queryset):
        if not request.user.is_superuser and request.user.role != 'judge':
            messages.error(request, '只有裁判和管理员可以确认成绩')
            return
        
        updated = queryset.filter(is_quit=False, confirmed=False).update(confirmed=True)
        if updated:
            messages.success(request, f'已确认 {updated} 条成绩记录')
        else:
            messages.warning(request, '没有可确认的成绩记录')
    confirm_selected.short_description = '确认选中成绩'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'athlete',
            'event'
        ).prefetch_related(
            'athlete__athlete_profile',
            'athlete__athlete_profile__team'
        )
    
    def has_change_permission(self, request, obj=None):
        if not obj:
            return True
        if request.user.is_superuser:
            return True
        if request.user.role == 'judge' and obj.event.judges.filter(id=request.user.id).exists():
            return not obj.confirmed
        return False

    # def export_as_excel(self, request, queryset):
    #     if not queryset.exists():
    #         self.message_user(request, '没有选中任何成绩记录，无法导出 Excel。')
    #         return HttpResponseRedirect(request.path_info)

    #     try:
    #         # 创建一个新的 Excel 工作簿和工作表
    #         workbook = openpyxl.Workbook()
    #         worksheet = workbook.active
    #         worksheet.title = '比赛成绩报告'

    #         # 添加表头
    #         headers = ['运动员', '项目', '成绩', '排名', '状态']
    #         worksheet.append(headers)

    #         # 填充数据
    #         for result in queryset:
    #             status = '已确认' if result.confirmed else '未确认'
    #             if result.is_quit:
    #                 status = '已弃权'
    #             row = [
    #                 result.athlete.username,
    #                 result.event.name,
    #                 str(result.score) if result.score else '-',
    #                 str(result.rank) if result.rank else '-',
    #                 status
    #             ]
    #             worksheet.append(row)

    #         # 本地测试
    #         # workbook.save('test_results.xlsx')

    #         # 创建响应对象
    #         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #         response['Content-Disposition'] = 'attachment; filename="results.xlsx"'
    #         # 禁止缓存
    #         response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    #         response['Pragma'] = 'no-cache'
    #         response['Expires'] = '0'

    #         # 将工作簿内容写入响应
    #         from openpyxl.writer.excel import save_virtual_workbook
    #         response.write(save_virtual_workbook(workbook))

    #         return response
    #     except Exception as e:
    #         self.message_user(request, f'导出 Excel 时出错：{str(e)}')
    #         return HttpResponseRedirect(request.path_info)

    # export_as_excel.short_description = '导出选中成绩为 Excel'

    def export_as_csv(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, '没有选中任何成绩记录，无法导出 CSV。')
            return HttpResponseRedirect(request.path_info)

        try:
            # 创建响应对象
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="results.csv"'
            # 禁止缓存
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'

            # 创建 CSV 写入器
            writer = csv.writer(response)

            # 添加表头
            headers = ['运动员', '项目', '成绩', '排名', '状态']
            writer.writerow(headers)

            # 填充数据
            for result in queryset:
                status = '已确认' if result.confirmed else '未确认'
                if result.is_quit:
                    status = '已弃权'
                row = [
                    result.athlete.username,
                    result.event.name,
                    str(result.score) if result.score else '-',
                    str(result.rank) if result.rank else '-',
                    status
                ]
                writer.writerow(row)

            return response
        except Exception as e:
            self.message_user(request, f'导出 CSV 时出错：{str(e)}')
            return HttpResponseRedirect(request.path_info)

    export_as_csv.short_description = '导出成绩'

    def generate_statistics(self, request, queryset):
        if not queryset.exists():
            self.message_user(request, '没有选中任何成绩记录，无法生成统计图表。')
            return HttpResponseRedirect(request.path_info)

        # 指定字体路径
        font_path = fm.findfont(fm.FontProperties(family='SimSun'))
        prop = fm.FontProperties(fname=font_path)

        # 设置全局字体
        plt.rcParams['font.family'] = prop.get_name()

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        try:
            # 项目平均成绩统计
            event_stats = queryset.values('event__name').annotate(
                avg_score=Avg('score'),
                total_athletes=Count('athlete'),
                quit_count=Count('id', filter=Q(is_quit=True))
            ).filter(score__isnull=False)

            if event_stats:
                events = [stat['event__name'] for stat in event_stats]
                scores = [stat['avg_score'] for stat in event_stats]

                rects = ax1.bar(events, scores)
                ax1.set_title('各项目平均成绩', fontproperties=prop)
                ax1.set_xlabel('项目名称', fontproperties=prop)
                ax1.set_ylabel('平均成绩', fontproperties=prop)
                ax1.set_xticklabels(events, rotation=45, fontproperties=prop)

                # 在项目平均成绩柱子上添加平均成绩数值标注
                def autolabel_event(rects):
                    for rect in rects:
                        height = rect.get_height()
                        ax1.annotate('{:.2f}'.format(height),
                                     xy=(rect.get_x() + rect.get_width() / 2, height),
                                     xytext=(0, 3),  # 3 points vertical offset
                                     textcoords="offset points",
                                     ha='center', va='bottom', fontproperties=prop)

                autolabel_event(rects)
            else:
                self.message_user(request, '没有符合条件的项目成绩数据，无法生成项目平均成绩图表。')

            # 团队成绩统计
            team_stats = Team.objects.annotate(
                medal_count=Count('athletes__user__results',
                                  filter=Q(athletes__user__results__rank__in=[1, 2, 3]))
            ).order_by('-score')[:10]

            if team_stats:
                teams = [team.name for team in team_stats]
                scores = [team.score for team in team_stats]
                medals = [team.medal_count for team in team_stats]

                x = range(len(teams))
                width = 0.35

                # 绘制总分柱子
                rects1 = ax2.bar([i - width/2 for i in x], scores, width, label='总分')
                # 绘制奖牌数柱子
                rects2 = ax2.bar([i + width/2 for i in x], medals, width, label='奖牌数')

                ax2.set_title('团队成绩统计', fontproperties=prop)
                ax2.set_xlabel('团队名称', fontproperties=prop)
                ax2.set_ylabel('分数/奖牌数', fontproperties=prop)
                ax2.set_xticks(x)
                ax2.set_xticklabels(teams, rotation=45, fontproperties=prop)
                ax2.legend(prop=prop)

                # 在柱子上添加总数标记
                def autolabel(rects):
                    for rect in rects:
                        height = rect.get_height()
                        ax2.annotate('{}'.format(height),
                                     xy=(rect.get_x() + rect.get_width() / 2, height),
                                     xytext=(0, 3),  # 3 points vertical offset
                                     textcoords="offset points",
                                     ha='center', va='bottom', fontproperties=prop)

                autolabel(rects1)
                autolabel(rects2)

            else:
                self.message_user(request, '没有符合条件的团队成绩数据，无法生成团队成绩图表。')

            plt.tight_layout()

            # 将图表转换为 base64 字符串
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            buffer.seek(0)
            image_png = buffer.getvalue()
            buffer.close()

            # 创建响应，明确指定字符编码
            response = HttpResponse(content_type='text/html; charset=utf-8')

            # 添加 CSS 样式
            css = """
            <style>
                h2 {
                    text-align: center;
                }
                img {
                    max-width: 100%;
                    height: auto;
                }
            </style>
            """
            response.write(css)

            response.write('<h2>成绩统计分析</h2>')
            response.write('<img src="data:image/png;base64,{}"/>'.format(
                base64.b64encode(image_png).decode()
            ))
            return response
        except Exception as e:
            self.message_user(request, f'生成统计图表时出错：{str(e)}')
            return HttpResponseRedirect(request.path_info)

    generate_statistics.short_description = "生成统计图表"

    def mark_as_quit(self, request, queryset):
        updated_count = queryset.update(is_quit=True)
        self.message_user(request, f'已将{updated_count}条记录标记为弃权')

    mark_as_quit.short_description = '标记为弃权'

    def remove_quit(self, request, queryset):
        updated_count = queryset.update(is_quit=False)
        self.message_user(request, f'已将{updated_count}条记录取消弃权')

    remove_quit.short_description = '取消弃权'

class TeamAdmin(admin.ModelAdmin):
    list_display = ('name', 'score', 'get_athletes_count', 'get_total_medals')
    search_fields = ('name',)
    
    def get_athletes_count(self, obj):
        return Athlete.objects.filter(team=obj).count()
    get_athletes_count.short_description = '运动员数量'
    
    def get_total_medals(self, obj):
        return Result.objects.filter(
            athlete__athlete_profile__team=obj,
            rank__in=[1, 2, 3]
        ).count()
    get_total_medals.short_description = '奖牌数'
    
    def get_active_athletes(self, obj):
        return obj.athletes.filter(is_active=True).count()
    get_active_athletes.short_description = '在籍运动员数'
    
    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'score')
        }),
        ('统计信息', {
            'fields': ('get_athletes_count', 'get_total_medals', 'get_active_athletes'),
            'classes': ('collapse',)
        })
    )
    readonly_fields = ('get_athletes_count', 'get_total_medals', 'get_active_athletes')
    
    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related('athletes')
    
    def has_delete_permission(self, request, obj=None):
        if obj and obj.athletes.exists():
            return False
        return super().has_delete_permission(request, obj)
    
    def save_model(self, request, obj, form, change):
        try:
            super().save_model(request, obj, form, change)
        except Exception as e:
            messages.error(request, f'保存失败：{str(e)}')

# 使用自定义admin站点注册模型
admin_site.register(User, CustomUserAdmin)
admin_site.register(Judge, JudgeAdmin)
admin_site.register(Athlete, AthleteAdmin)
admin_site.register(Event, EventAdmin)
admin_site.register(Result, ResultAdmin)
admin_site.register(Team, TeamAdmin)

# 自定义管理后台标题
admin_site.site_header = '田径运动会管理系统'
admin_site.site_title = '田径运动会管理系统'
admin_site.index_title = '管理面板'
